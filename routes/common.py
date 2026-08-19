"""路由共享工具模块 — Excel读取、PDF渲染、安全校验等通用函数

所有 Blueprint 共享的纯工具函数集中在此，避免循环导入。
依赖 Flask current_app 的地方在函数内部使用 current_app。
"""
import os
import re
import gc
import logging
import traceback
import threading
from datetime import datetime, timezone, timedelta

import db

logger = logging.getLogger(__name__)

_CST = timezone(timedelta(hours=8))

# 从 excel_analyzers 导入内存日志工具（ExcelReader 依赖）
try:
    from excel_analyzers import _log_mem
except ImportError:
    def _log_mem(msg):
        pass


# ==================== Excel 读取器 ====================

class ExcelReader:
    """统一的 Excel 读取器，支持 .xls、.xlsx、.csv 格式，以及 HTML 格式的 Excel 文件"""

    def __init__(self, file_path):
        self.file_path = file_path
        self.ext = os.path.splitext(file_path)[1].lower()
        self._wb = None
        self._is_xls = self.ext == '.xls'
        self._is_csv = self.ext == '.csv'
        self._is_html = False
        self._read_only = False
        self._csv_data = None  # CSV 数据缓存

    def open(self):
        # CSV 格式：用 pandas 读取（自动检测编码）
        if self._is_csv:
            import pandas as pd
            import csv as csv_module
            # 自动检测分隔符
            try:
                with open(self.file_path, 'r', encoding='utf-8', errors='ignore') as f:
                    sample = f.read(8192)
                try:
                    dialect = csv_module.Sniffer().sniff(sample, delimiters=',;\t|')
                    delimiter = dialect.delimiter
                except csv_module.Error:
                    delimiter = ','
            except Exception:
                delimiter = ','

            # 尝试多种编码 + 自动检测的分隔符
            for encoding in ['utf-8-sig', 'utf-8', 'gbk', 'gb2312', 'latin1']:
                try:
                    df = pd.read_csv(
                        self.file_path,
                        encoding=encoding,
                        dtype=str,
                        keep_default_na=False,
                        sep=delimiter,
                        engine='python',
                        on_bad_lines='skip'
                    )
                    if len(df.columns) > 1 or (len(df.columns) == 1 and delimiter != ','):
                        self._csv_data = df
                        return self
                    # 如果只有1列且分隔符是逗号，可能是分隔符不对，尝试其他分隔符
                    for alt_delim in [';', '\t', '|']:
                        try:
                            df2 = pd.read_csv(
                                self.file_path,
                                encoding=encoding,
                                dtype=str,
                                keep_default_na=False,
                                sep=alt_delim,
                                engine='python',
                                on_bad_lines='skip'
                            )
                            if len(df2.columns) > 1:
                                self._csv_data = df2
                                return self
                        except Exception:
                            continue
                except Exception:
                    continue

            # 全部失败，用 csv 模块兜底
            try:
                for encoding in ['utf-8-sig', 'utf-8', 'gbk', 'latin1']:
                    try:
                        rows = []
                        with open(self.file_path, 'r', encoding=encoding, errors='replace', newline='') as f:
                            reader = csv_module.reader(f)
                            for row in reader:
                                rows.append(row)
                        if rows and len(rows) > 0:
                            import pandas as pd
                            df = pd.DataFrame(rows[1:], columns=rows[0]) if len(rows) > 1 else pd.DataFrame(columns=rows[0])
                            self._csv_data = df
                            return self
                    except Exception:
                        continue
            except Exception:
                pass

            raise ValueError('无法解析CSV文件，请检查文件格式和编码')

        # 检测是否是 HTML 格式的假 Excel 文件
        if self._is_xls:
            with open(self.file_path, 'rb') as f:
                header = f.read(100)
            if b'<html' in header.lower() or b'<!doctype' in header.lower():
                self._is_html = True

        if self._is_html:
            return self  # HTML 格式不需要打开

        if self._is_xls:
            import xlrd
            try:
                self._wb = xlrd.open_workbook(self.file_path)
            except Exception as e:
                raise ValueError(f'无法读取 .xls 文件: {str(e)}。如果文件是从网页下载的，请转换为 .xlsx 格式后再上传。')
        else:
            from openpyxl import load_workbook
            # 先尝试 read_only=True（内存效率高），如果读到的行太少则回退到 read_only=False
            self._wb = load_workbook(self.file_path, data_only=True, read_only=True)
            self._read_only = True
        return self

    def close(self):
        if self._is_csv:
            self._csv_data = None
            return
        if self._wb and not self._is_xls and not self._is_html:
            self._wb.close()

    @property
    def sheetnames(self):
        if self._is_csv:
            return ['Sheet1']
        if self._is_html:
            return ['Sheet1']  # HTML 格式默认返回一个 sheet
        if self._is_xls:
            return self._wb.sheet_names()
        return self._wb.sheetnames

    def get_sheet_data(self, sheet_name):
        """获取指定 sheet 的所有行数据，返回 list of lists"""
        if self._is_csv:
            if self._csv_data is None:
                self.open()
            df = self._csv_data
            try:
                # 表头 + 数据行（使用 values 提高性能）
                headers = [str(h).strip() if h is not None else '' for h in df.columns.tolist()]
                rows = [headers]
                # 使用 itertuples 比 iterrows 更快
                for row in df.itertuples(index=False, name=None):
                    rows.append([str(v).strip() if v is not None else '' for v in row])
                return rows
            except Exception as e:
                logger.warning(f"CSV get_sheet_data 失败，尝试备选方式: {e}")
                # 备选：直接转 list
                headers = [str(h) for h in df.columns.tolist()]
                rows = [headers] + df.values.tolist()
                return [[str(v).strip() if v is not None else '' for v in row] for row in rows]

        if self._is_html:
            return self._parse_html_excel()

        if self._is_xls:
            sheet = self._wb.sheet_by_name(sheet_name)
            rows = []
            for row_idx in range(sheet.nrows):
                row = [str(sheet.cell_value(row_idx, col_idx)).strip() if sheet.cell_value(row_idx, col_idx) != '' else ''
                       for col_idx in range(sheet.ncols)]
                rows.append(row)
            return rows
        else:
            ws = self._wb[sheet_name]
            rows = [[str(cell).strip() if cell is not None else '' for cell in row] for row in ws.iter_rows(values_only=True)]

            # read_only=True 模式下某些 Excel 文件只能读到1-2行（openpyxl已知bug）
            # 检测到行数异常少时，回退到 read_only=False 重新读取
            if self._read_only and len(rows) <= 2:
                try:
                    self._wb.close()
                except Exception:
                    pass
                from openpyxl import load_workbook
                self._wb = load_workbook(self.file_path, data_only=True, read_only=False)
                self._read_only = False
                ws = self._wb[sheet_name]
                rows = [[str(cell).strip() if cell is not None else '' for cell in row] for row in ws.iter_rows(values_only=True)]

            return rows

    def get_headers_only(self, sheet_name=None):
        """轻量级方法：只读取表头行，不加载全部数据（避免 OOM）"""
        if self._is_html:
            return self._parse_html_headers_only()

        # 非 HTML 格式：打开后只读第一行
        if not self._wb:
            self.open()
        rows = self.get_sheet_data(sheet_name or (self.sheetnames[0] if self.sheetnames else None))
        return rows[0] if rows else []

    def _parse_html_headers_only(self):
        """只解析 HTML 表头，不加载整个文件到内存（流式读取前 N KB）"""
        try:
            # 只读取文件前 500KB 来提取表头和第一行数据
            READ_SIZE = 500 * 1024
            with open(self.file_path, 'r', encoding='utf-8', errors='ignore') as f:
                chunk = f.read(READ_SIZE)

            # 用正则提取 <th> 标签内容作为表头
            th_pattern = re.compile(r'<th[^>]*>(.*?)</th>', re.IGNORECASE | re.DOTALL)
            th_matches = th_pattern.findall(chunk)

            # 清理 HTML 标签和实体
            def clean_html_text(text):
                text = re.sub(r'<[^>]+>', '', text)
                text = text.replace('&amp;', '&').replace('&lt;', '<').replace('&gt;', '>')
                text = text.replace('&quot;', '"').replace('&#39;', "'").replace('&nbsp;', ' ')
                text = re.sub(r'&#(\d+);', lambda m: chr(int(m.group(1))) if int(m.group(1)) < 65536 else '', text)
                text = re.sub(r'&#x([0-9a-fA-F]+);', lambda m: chr(int(m.group(1), 16)) if int(m.group(1), 16) < 65536 else '', text)
                return text.strip()

            all_headers = [clean_html_text(m) for m in th_matches if clean_html_text(m)]

            if not all_headers:
                # 尝试从第一个 <tr> 中的 <td> 提取
                tr_pattern = re.compile(r'<tr[^>]*>(.*?)</tr>', re.IGNORECASE | re.DOTALL)
                tr_matches = tr_pattern.findall(chunk)
                if tr_matches:
                    td_pattern = re.compile(r'<td[^>]*>(.*?)</td>', re.IGNORECASE | re.DOTALL)
                    td_matches = td_pattern.findall(tr_matches[0])
                    all_headers = [clean_html_text(m) for m in td_matches if clean_html_text(m)]

            if not all_headers:
                all_headers = ['Project', 'Key', 'Summary', 'Issue Type', 'Status',
                              'Priority', 'Resolution', 'Assignee', 'Reporter', 'Creator',
                              'Created', 'Last Viewed', 'Updated', 'Resolved', 'Affects Version/s']

            # 提取第一行数据（thead 之后的第一个 tr 中的 td）
            first_data_row = []
            thead_end = chunk.lower().find('</thead>')
            if thead_end >= 0:
                after_thead = chunk[thead_end:]
                tr_match = re.search(r'<tr[^>]*>(.*?)</tr>', after_thead, re.IGNORECASE | re.DOTALL)
                if tr_match:
                    td_pattern = re.compile(r'<td[^>]*>(.*?)</td>', re.IGNORECASE | re.DOTALL)
                    td_matches = td_pattern.findall(tr_match.group(1))
                    first_data_row = [clean_html_text(m) for m in td_matches]

            # 估算数据行数：统计文件中 <tr 标签出现次数
            data_row_count = 0
            try:
                with open(self.file_path, 'r', encoding='utf-8', errors='ignore') as f:
                    for piece in iter(lambda: f.read(1024 * 1024), ''):
                        data_row_count += piece.count('<tr')
            except Exception:
                pass
            data_row_count = max(0, data_row_count - 1)

            del chunk
            gc.collect()

            logger.info(f"HTML表头流式解析完成: {len(all_headers)}列, 约{data_row_count}行")
            return all_headers, first_data_row, data_row_count

        except Exception as e:
            raise ValueError(f'解析 HTML 格式 Excel 文件失败: {str(e)}')

    def _parse_html_excel(self):
        """使用 lxml 高性能解析 HTML 格式 Excel 文件"""
        try:
            from lxml import html as lxml_html
            import html as html_module

            with open(self.file_path, 'rb') as f:
                tree = lxml_html.fromstring(f.read())

            tables = tree.xpath('//table')
            if not tables:
                return self._parse_html_excel_regex()

            table = tables[0]
            if len(tables) > 1:
                max_rows = 0
                for t in tables:
                    cnt = len(t.xpath('.//tr'))
                    if cnt > max_rows:
                        max_rows = cnt
                        table = t

            # === 提取表头 ===
            all_headers = []
            thead = table.find('thead')
            if thead is not None:
                thead_rows = thead.findall('tr')
                if thead_rows:
                    best_row = max(thead_rows, key=lambda r: len(r.findall('th')))
                    for th in best_row.findall('th'):
                        text = html_module.unescape(lxml_html.tostring(th, encoding='unicode', method='text').strip())
                        colspan = th.get('colspan', '1')
                        try:
                            span = int(colspan)
                        except (ValueError, TypeError):
                            span = 1
                        for _ in range(span):
                            all_headers.append(text if text else '')
                else:
                    for th in thead.iter('th'):
                        text = html_module.unescape(lxml_html.tostring(th, encoding='unicode', method='text').strip())
                        if text:
                            all_headers.append(text)
            if not all_headers:
                first_tr = table.find('.//tr')
                if first_tr is not None:
                    for td in first_tr.iter('td'):
                        text = html_module.unescape(lxml_html.tostring(td, encoding='unicode', method='text').strip())
                        if text:
                            all_headers.append(text)

            if not all_headers:
                all_headers = ['Project', 'Key', 'Summary', 'Issue Type', 'Status',
                              'Priority', 'Resolution', 'Assignee', 'Reporter', 'Creator',
                              'Created', 'Last Viewed', 'Updated', 'Resolved', 'Affects Version/s']

            full_headers = list(all_headers)
            keep_cols_sorted = list(range(len(all_headers)))
            _log_mem(f"HTML lxml解析：保留全部{len(full_headers)}列")

            # === 提取数据行 ===
            result_rows = [full_headers]
            row_count = 0

            tbody = table.find('tbody')
            tr_source = tbody if tbody is not None else table

            for tr in tr_source.iter('tr'):
                if tr.find('th') is not None:
                    continue
                tds = tr.findall('td')
                if not tds:
                    continue
                td_texts = [html_module.unescape(td.text_content().strip()) for td in tds]
                row = []
                for col_idx in keep_cols_sorted:
                    if col_idx < len(td_texts):
                        row.append(td_texts[col_idx])
                    else:
                        row.append('')
                if any(c.strip() for c in row):
                    result_rows.append(row)
                    row_count += 1

            del tree, table
            gc.collect()

            logger.info(f"HTML lxml解析完成：{row_count}行 x {len(full_headers)}列")
            _log_mem(f"HTML lxml解析完成：{row_count}行 x {len(full_headers)}列")
            return result_rows

        except Exception as e:
            error_msg = f'解析 HTML 格式 Excel 文件失败: {type(e).__name__}: {str(e)}'
            logger.error(error_msg)
            logger.error(traceback.format_exc())
            logger.warning("lxml 解析失败，回退到正则解析")
            return self._parse_html_excel_regex()

    def _parse_html_excel_regex(self):
        """正则方式解析 HTML 格式（回退方案）"""
        try:
            def clean_html_text(text):
                text = re.sub(r'<[^>]+>', '', text)
                text = text.replace('&amp;', '&').replace('&lt;', '<').replace('&gt;', '>')
                text = text.replace('&quot;', '"').replace('&#39;', "'").replace('&nbsp;', ' ')
                text = re.sub(r'&#(\d+);', lambda m: chr(int(m.group(1))) if int(m.group(1)) < 65536 else '', text)
                text = re.sub(r'&#x([0-9a-fA-F]+);', lambda m: chr(int(m.group(1), 16)) if int(m.group(1), 16) < 65536 else '', text)
                return text.strip()

            READ_SIZE = 1024 * 1024
            with open(self.file_path, 'r', encoding='utf-8', errors='ignore') as f:
                head_chunk = f.read(READ_SIZE)

            th_full_pattern = re.compile(r'<th([^>]*)>(.*?)</th>', re.IGNORECASE | re.DOTALL)
            th_full_matches = th_full_pattern.findall(head_chunk)
            all_headers = []
            for attrs_str, content in th_full_matches:
                text = clean_html_text(content)
                colspan_match = re.search(r'colspan\s*=\s*["\']?(\d+)', attrs_str, re.IGNORECASE)
                span = int(colspan_match.group(1)) if colspan_match else 1
                for _ in range(span):
                    all_headers.append(text)

            if not all_headers:
                tr_pattern = re.compile(r'<tr[^>]*>(.*?)</tr>', re.IGNORECASE | re.DOTALL)
                tr_match = tr_pattern.search(head_chunk)
                if tr_match:
                    td_pattern = re.compile(r'<td[^>]*>(.*?)</td>', re.IGNORECASE | re.DOTALL)
                    td_matches = td_pattern.findall(tr_match.group(1))
                    all_headers = [clean_html_text(m) for m in td_matches if clean_html_text(m)]

            if not all_headers:
                all_headers = ['Project', 'Key', 'Summary', 'Issue Type', 'Status',
                              'Priority', 'Resolution', 'Assignee', 'Reporter', 'Creator',
                              'Created', 'Last Viewed', 'Updated', 'Resolved', 'Affects Version/s']

            full_headers = list(all_headers)
            keep_cols_sorted = list(range(len(all_headers)))
            _log_mem(f"HTML正则解析：保留全部{len(full_headers)}列")

            result_rows = [full_headers]
            row_count = 0
            td_pattern = re.compile(r'<td[^>]*>(.*?)</td>', re.IGNORECASE | re.DOTALL)

            thead_end_pos = head_chunk.lower().find('</thead>')
            skip_header = thead_end_pos >= 0
            table_start_pos = head_chunk.lower().find('<table')
            if not skip_header and table_start_pos >= 0:
                start_pos = table_start_pos
            elif skip_header:
                start_pos = thead_end_pos + 8
            else:
                start_pos = 0

            table_ended = False
            with open(self.file_path, 'r', encoding='utf-8', errors='ignore') as f:
                f.seek(start_pos)
                buffer = ''
                CHUNK_SIZE = 512 * 1024

                while True and not table_ended:
                    piece = f.read(CHUNK_SIZE)
                    if not piece:
                        break
                    buffer += piece

                    while True:
                        tr_start = buffer.find('<tr')
                        if tr_start == -1:
                            buffer = ''
                            break
                        tr_end = buffer.find('</tr>', tr_start)
                        if tr_end == -1:
                            buffer = buffer[tr_start:]
                            if len(buffer) > 1024 * 1024:
                                buffer = ''
                            break

                        tr_content = buffer[tr_start:tr_end + 5]
                        buffer = buffer[tr_end + 5:]

                        if '<th' in tr_content.lower():
                            next_tr = buffer.find('<tr')
                            between = buffer[:next_tr] if next_tr >= 0 else buffer
                            if '</table>' in between.lower():
                                table_ended = True
                                break
                            continue

                        td_matches = td_pattern.findall(tr_content)
                        if not td_matches:
                            continue

                        row = []
                        for col_idx in keep_cols_sorted:
                            if col_idx < len(td_matches):
                                row.append(clean_html_text(td_matches[col_idx]))
                            else:
                                row.append('')

                        if any(c.strip() for c in row):
                            result_rows.append(row)
                            row_count += 1

                        next_tr = buffer.find('<tr')
                        between = buffer[:next_tr] if next_tr >= 0 else buffer
                        if '</table>' in between.lower():
                            table_ended = True
                            break

                    del piece

            del buffer, head_chunk
            gc.collect()

            logger.info(f"HTML正则解析完成：{row_count}行 x {len(full_headers)}列, table_ended={table_ended}")
            _log_mem(f"HTML正则解析完成：{row_count}行 x {len(full_headers)}列")
            return result_rows

        except Exception as e:
            error_msg = f'解析 HTML 格式 Excel 文件失败: {type(e).__name__}: {str(e)}'
            logger.error(error_msg)
            logger.error(traceback.format_exc())
            raise ValueError(error_msg)

    def get_sheet_names(self):
        """获取所有 sheet 名称"""
        return self.sheetnames


def read_excel_file(file_path, sheet_name=None):
    """读取 Excel 文件，返回 (sheet_names, sheet_data) 或 (sheet_names, None)"""
    reader = ExcelReader(file_path)
    reader.open()
    sheet_names = reader.get_sheet_names()
    sheet_data = None
    if sheet_name:
        sheet_data = reader.get_sheet_data(sheet_name)
    reader.close()
    return sheet_names, sheet_data


# ==================== 安全校验 ====================

def validate_file_id(file_id):
    """验证 file_id 格式，防止路径遍历攻击
    合法格式：16位十六进制字符（md5 hexdigest[:16]）
    """
    if not file_id or not isinstance(file_id, str):
        return False
    return bool(re.match(r'^[a-f0-9]{16}$', file_id))


# ==================== PDF 渲染 ====================

try:
    from playwright.sync_api import sync_playwright as _sync_playwright
except ImportError:
    _sync_playwright = None

_pdf_render_lock = threading.Lock()
_pw_instance = None
_pw_browser = None

# MD2PDF 预览样式 CSS
MD2PDF_PREVIEW_CSS = """
        body { font-family: -apple-system, "Segoe UI", Roboto, "PingFang SC", "Microsoft YaHei", sans-serif; padding: 40px; background: #fff; color: #333; font-size: 14px; line-height: 1.8; }
        h1 { font-size: 26px; font-weight: bold; border-bottom: 2px solid #667eea; padding-bottom: 12px; margin-bottom: 20px; color: #333; }
        h2 { font-size: 22px; font-weight: bold; margin-top: 30px; margin-bottom: 15px; color: #333; }
        h3 { font-size: 19px; font-weight: bold; margin-top: 25px; margin-bottom: 12px; color: #333; }
        h4 { font-size: 17px; font-weight: bold; margin-top: 20px; margin-bottom: 10px; color: #333; }
        p { margin-bottom: 15px; text-indent: 2em; }
        ul, ol { margin-bottom: 15px; padding-left: 2em; }
        li { margin-bottom: 8px; }
        blockquote { border-left: 4px solid #667eea; padding: 12px 16px; margin: 15px 0; color: #666; font-style: italic; background-color: #f5f3ff; border-radius: 0 8px 8px 0; }
        code { background-color: #f4f4f4; padding: 3px 8px; border-radius: 4px; font-family: "Consolas", "Courier New", monospace; font-size: 0.9em; }
        pre { background-color: #2d2d2d; color: #ccc; padding: 16px; border-radius: 10px; overflow-x: auto; margin-bottom: 15px; white-space: pre-wrap; word-break: break-all; }
        pre code { background-color: transparent; padding: 0; color: inherit; }
        table { width: 100%; border-collapse: collapse; margin-bottom: 20px; font-size: 13px; }
        th, td { border: 1px solid #ddd; padding: 10px; text-align: left; }
        th { background-color: #f5f3ff; font-weight: bold; color: #667eea; }
        a { color: #667eea; text-decoration: none; font-weight: 500; }
        img { max-width: 100%; height: auto; border-radius: 8px; }
        hr { border: none; border-top: 2px dashed #ddd; margin: 25px 0; }
"""


def get_pw_browser():
    """获取或创建全局 Chromium 浏览器实例（复用，避免每次冷启动）"""
    global _pw_instance, _pw_browser
    if _sync_playwright is None:
        raise RuntimeError("Playwright 未安装，请运行: pip install playwright && playwright install chromium")
    if _pw_browser is not None:
        try:
            _ = _pw_browser.version
            return _pw_browser
        except Exception:
            _pw_browser = None
            logger.warning("Chromium 浏览器已断开，正在重新创建...")
    if _pw_instance is None:
        _pw_instance = _sync_playwright().start()
    try:
        _pw_browser = _pw_instance.chromium.launch(headless=True, channel="chrome")
    except Exception:
        _pw_browser = _pw_instance.chromium.launch(headless=True)
    logger.info("Chromium 浏览器实例已创建（全局复用）")
    return _pw_browser


def render_pdf(html_path, pdf_path, margin=None, extra_wait_ms=0, wait_selector=None):
    """使用全局 Chromium 实例渲染 PDF（线程安全）"""
    if margin is None:
        margin = {'top': '2cm', 'right': '2cm', 'bottom': '2cm', 'left': '2cm'}
    with _pdf_render_lock:
        browser = get_pw_browser()
        context = browser.new_context()
        page = context.new_page()
        try:
            page.goto(f'file://{html_path}')
            page.wait_for_load_state('networkidle')
            if wait_selector:
                try:
                    page.wait_for_selector(wait_selector, timeout=5000)
                except Exception:
                    pass
            if extra_wait_ms > 0:
                page.wait_for_timeout(extra_wait_ms)
            page.pdf(
                path=pdf_path,
                format='A4',
                margin=margin,
                print_background=True
            )
        finally:
            context.close()


# ==================== 后台任务共享状态 ====================
# 全局后台任务字典（所有 Blueprint 共享）
background_tasks = {}


def load_task_meta(task_id):
    """从数据库加载任务元数据"""
    task = db.get_task(task_id)
    if not task:
        return None
    return {
        'task_id': task_id,
        'task_type': task.get('task_type', ''),
        'status': task.get('status', 'pending'),
        'progress': task.get('progress', 0),
        'result': task.get('result'),
        'error': task.get('error'),
    }


def save_task_meta(task_id, task_data):
    """持久化任务元数据到数据库"""
    status = task_data.get('status', 'pending')
    progress = task_data.get('progress', 0)
    result = task_data.get('result')
    error = task_data.get('error')
    existing = db.get_task(task_id)
    if existing:
        db.update_task(task_id, status=status, progress=progress, result=result, error=error)
    else:
        db.create_task(task_id, task_data.get('task_type', 'unknown'))
        db.update_task(task_id, status=status, progress=progress, result=result, error=error)


def delete_task_meta(task_id):
    """删除任务元数据"""
    db.delete_task(task_id)
